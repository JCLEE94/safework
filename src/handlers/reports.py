from typing import Optional, Dict, Any, List
from datetime import datetime, timedelta
from fastapi import APIRouter, Depends, HTTPException, Response
from fastapi.responses import StreamingResponse
from sqlalchemy import select, func, and_, or_
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload
import io
import json
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from reportlab.lib.pagesizes import A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from reportlab.lib.units import inch
from reportlab.pdfbase import pdfutils
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.pdfbase import pdfmetrics
import os

from src.config.database import get_db
from src.models import (
    Worker, HealthExam, WorkEnvironment, HealthEducation, 
    ChemicalSubstance, AccidentReport
)
from src.schemas.base import BaseResponse

router = APIRouter(prefix="/api/v1/reports", tags=["reports"])

# 한글 폰트 등록 (PDF용)
try:
    font_path = "/usr/share/fonts/truetype/nanum/NanumGothic.ttf"
    if os.path.exists(font_path):
        pdfmetrics.registerFont(TTFont('NanumGothic', font_path))
except:
    pass

class ReportFilters:
    def __init__(self, data: Dict[str, Any]):
        self.start_date = data.get('startDate')
        self.end_date = data.get('endDate') 
        self.department = data.get('department')
        self.location = data.get('location')
        self.status = data.get('status')

class ReportGenerator:
    def __init__(self, db: AsyncSession):
        self.db = db
    
    async def generate_worker_health_summary(self, filters: ReportFilters) -> bytes:
        """근로자 건강현황 종합보고서 (Excel)"""
        # 데이터 조회
        query = select(Worker).options(selectinload(Worker.health_exams))
        if filters.department:
            query = query.where(Worker.department.contains(filters.department))
        
        result = await self.db.execute(query)
        workers = result.scalars().all()
        
        # Excel 생성
        wb = Workbook()
        ws = wb.active
        ws.title = "근로자 건강현황"
        
        # 헤더 스타일
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # 헤더
        headers = ["번호", "성명", "부서", "직위", "입사일", "건강상태", "최근검진일", "차기검진일"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # 데이터 입력
        for row, worker in enumerate(workers, 2):
            latest_exam = None
            if worker.health_exams:
                latest_exam = max(worker.health_exams, key=lambda x: x.exam_date)
            
            ws.cell(row=row, column=1, value=row-1)
            ws.cell(row=row, column=2, value=worker.name)
            ws.cell(row=row, column=3, value=worker.department)
            ws.cell(row=row, column=4, value=worker.position)
            ws.cell(row=row, column=5, value=worker.hire_date.strftime('%Y-%m-%d') if worker.hire_date else '')
            ws.cell(row=row, column=6, value=worker.health_status.value if worker.health_status else '')
            
            if latest_exam:
                ws.cell(row=row, column=7, value=latest_exam.exam_date.strftime('%Y-%m-%d'))
                # 차기 검진일 (1년 후)
                next_exam = latest_exam.exam_date + timedelta(days=365)
                ws.cell(row=row, column=8, value=next_exam.strftime('%Y-%m-%d'))
        
        # 열 너비 조정
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[chr(64 + col)].width = 15
        
        # 통계 시트 추가
        stats_ws = wb.create_sheet("통계")
        
        # 건강상태별 통계
        from collections import Counter
        health_stats = Counter([w.health_status.value if w.health_status else '미정' for w in workers])
        
        stats_ws.cell(row=1, column=1, value="건강상태별 통계").font = Font(bold=True, size=14)
        for row, (status, count) in enumerate(health_stats.items(), 3):
            stats_ws.cell(row=row, column=1, value=status)
            stats_ws.cell(row=row, column=2, value=count)
        
        # 파일 저장
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()
    
    async def generate_work_env_compliance(self, filters: ReportFilters) -> bytes:
        """작업환경측정 법규준수 보고서 (PDF)"""
        # 데이터 조회
        query = select(WorkEnvironment)
        conditions = []
        
        if filters.start_date:
            conditions.append(WorkEnvironment.measurement_date >= datetime.fromisoformat(filters.start_date))
        if filters.end_date:
            conditions.append(WorkEnvironment.measurement_date <= datetime.fromisoformat(filters.end_date))
        if filters.location:
            conditions.append(WorkEnvironment.location.contains(filters.location))
        
        if conditions:
            query = query.where(and_(*conditions))
        
        result = await self.db.execute(query)
        measurements = result.scalars().all()
        
        # PDF 생성
        output = io.BytesIO()
        doc = SimpleDocTemplate(output, pagesize=A4)
        elements = []
        
        # 스타일 설정
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontName='NanumGothic' if 'NanumGothic' in pdfmetrics.getRegisteredFontNames() else 'Helvetica-Bold',
            fontSize=16,
            spaceAfter=30,
            alignment=1  # 중앙 정렬
        )
        
        # 제목
        title = Paragraph("작업환경측정 법규준수 보고서", title_style)
        elements.append(title)
        elements.append(Spacer(1, 20))
        
        # 요약 정보
        total_measurements = len(measurements)
        pass_count = len([m for m in measurements if m.result.value == '적합'])
        fail_count = len([m for m in measurements if m.result.value == '부적합'])
        
        summary_data = [
            ['항목', '값'],
            ['총 측정 건수', str(total_measurements)],
            ['적합 건수', str(pass_count)],
            ['부적합 건수', str(fail_count)],
            ['적합률', f"{(pass_count/total_measurements*100):.1f}%" if total_measurements > 0 else "0%"]
        ]
        
        summary_table = Table(summary_data)
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        elements.append(summary_table)
        elements.append(Spacer(1, 30))
        
        # 상세 측정 결과
        if measurements:
            detail_title = Paragraph("상세 측정 결과", styles['Heading2'])
            elements.append(detail_title)
            elements.append(Spacer(1, 10))
            
            detail_data = [['장소', '측정항목', '측정일', '측정값', '기준값', '결과']]
            for m in measurements[:20]:  # 최대 20개만 표시
                detail_data.append([
                    m.location,
                    m.measurement_type.value,
                    m.measurement_date.strftime('%Y-%m-%d'),
                    str(m.measured_value),
                    str(m.standard_value),
                    m.result.value
                ])
            
            detail_table = Table(detail_data)
            detail_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            
            elements.append(detail_table)
        
        doc.build(elements)
        output.seek(0)
        return output.getvalue()
    
    async def generate_monthly_dashboard(self, filters: ReportFilters) -> bytes:
        """월간 대시보드 보고서 (PDF)"""
        # 각종 통계 데이터 수집
        start_date = datetime.fromisoformat(filters.start_date) if filters.start_date else datetime.now().replace(day=1)
        end_date = datetime.fromisoformat(filters.end_date) if filters.end_date else datetime.now()
        
        # 근로자 통계
        worker_count = await self.db.scalar(select(func.count(Worker.id)))
        
        # 건강검진 통계
        health_exam_query = select(func.count(HealthExam.id)).where(
            and_(HealthExam.exam_date >= start_date, HealthExam.exam_date <= end_date)
        )
        health_exam_count = await self.db.scalar(health_exam_query)
        
        # 작업환경측정 통계
        work_env_query = select(func.count(WorkEnvironment.id)).where(
            and_(WorkEnvironment.measurement_date >= start_date, WorkEnvironment.measurement_date <= end_date)
        )
        work_env_count = await self.db.scalar(work_env_query)
        
        # 사고 통계
        accident_query = select(func.count(AccidentReport.id)).where(
            and_(AccidentReport.accident_date >= start_date, AccidentReport.accident_date <= end_date)
        )
        accident_count = await self.db.scalar(accident_query)
        
        # PDF 생성
        output = io.BytesIO()
        doc = SimpleDocTemplate(output, pagesize=A4)
        elements = []
        
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontName='NanumGothic' if 'NanumGothic' in pdfmetrics.getRegisteredFontNames() else 'Helvetica-Bold',
            fontSize=18,
            spaceAfter=30,
            alignment=1
        )
        
        # 제목
        period = f"{start_date.strftime('%Y년 %m월')} ~ {end_date.strftime('%Y년 %m월')}"
        title = Paragraph(f"월간 대시보드 보고서<br/>{period}", title_style)
        elements.append(title)
        elements.append(Spacer(1, 30))
        
        # 주요 지표
        dashboard_data = [
            ['구분', '건수', '비고'],
            ['등록 근로자', str(worker_count), '전체'],
            ['건강검진', str(health_exam_count), '기간 내'],
            ['작업환경측정', str(work_env_count), '기간 내'],
            ['산업재해', str(accident_count), '기간 내']
        ]
        
        dashboard_table = Table(dashboard_data)
        dashboard_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.navy),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 14),
            ('FONTSIZE', (0, 1), (-1, -1), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.lightblue),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))
        
        elements.append(dashboard_table)
        elements.append(Spacer(1, 30))
        
        # 보고서 생성 정보
        generated_time = datetime.now().strftime('%Y년 %m월 %d일 %H시 %M분')
        footer = Paragraph(f"생성일시: {generated_time}", styles['Normal'])
        elements.append(footer)
        
        doc.build(elements)
        output.seek(0)
        return output.getvalue()

@router.post("/generate")
async def generate_report(
    request: Dict[str, Any],
    db: AsyncSession = Depends(get_db)
):
    """보고서 생성"""
    template_id = request.get("template_id")
    filters = ReportFilters(request.get("filters", {}))
    
    if not template_id:
        raise HTTPException(status_code=400, detail="템플릿 ID가 필요합니다")
    
    generator = ReportGenerator(db)
    
    try:
        if template_id == "worker_health_summary":
            content = await generator.generate_worker_health_summary(filters)
            filename = f"근로자_건강현황_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            
        elif template_id == "work_env_compliance":
            content = await generator.generate_work_env_compliance(filters)
            filename = f"작업환경측정_법규준수_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
            media_type = "application/pdf"
            
        elif template_id == "monthly_dashboard":
            content = await generator.generate_monthly_dashboard(filters)
            filename = f"월간_대시보드_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
            media_type = "application/pdf"
            
        else:
            # 기타 템플릿은 간단한 텍스트 보고서로 생성
            content = f"보고서 템플릿 '{template_id}'는 아직 구현되지 않았습니다.\n생성 요청: {datetime.now()}".encode('utf-8')
            filename = f"보고서_{template_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.txt"
            media_type = "text/plain"
        
        return Response(
            content=content,
            media_type=media_type,
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"보고서 생성 실패: {str(e)}")

@router.get("/recent")
async def get_recent_reports():
    """최근 생성된 보고서 목록"""
    # 실제로는 데이터베이스에서 보고서 생성 이력을 조회해야 함
    # 현재는 예시 데이터 반환
    recent_reports = [
        {
            "name": "근로자 건강현황 종합보고서",
            "type": "excel",
            "generated_at": "2025-06-26 13:20:00",
            "download_url": "/api/v1/reports/download/sample1"
        },
        {
            "name": "작업환경측정 법규준수 보고서", 
            "type": "pdf",
            "generated_at": "2025-06-25 15:30:00",
            "download_url": "/api/v1/reports/download/sample2"
        }
    ]
    
    return {"reports": recent_reports}

@router.get("/templates")
async def get_report_templates():
    """사용 가능한 보고서 템플릿 목록"""
    templates = [
        {
            "id": "worker_health_summary",
            "name": "근로자 건강현황 종합보고서",
            "description": "전체 근로자의 건강검진 현황 및 통계",
            "category": "건강관리",
            "type": "excel"
        },
        {
            "id": "work_env_compliance",
            "name": "작업환경측정 법규준수 보고서",
            "description": "작업환경측정 결과 및 법규 준수 현황",
            "category": "작업환경",
            "type": "pdf"
        },
        {
            "id": "monthly_dashboard",
            "name": "월간 대시보드 보고서",
            "description": "주요 지표별 월간 현황 종합",
            "category": "종합",
            "type": "pdf"
        }
    ]
    
    return {"templates": templates}

@router.get("/statistics")
async def get_report_statistics(db: AsyncSession = Depends(get_db)):
    """보고서 관련 통계"""
    # 각종 통계 데이터 조회
    worker_count = await db.scalar(select(func.count(Worker.id)))
    health_exam_count = await db.scalar(select(func.count(HealthExam.id)))
    work_env_count = await db.scalar(select(func.count(WorkEnvironment.id)))
    accident_count = await db.scalar(select(func.count(AccidentReport.id)))
    
    return {
        "total_workers": worker_count,
        "total_health_exams": health_exam_count,
        "total_work_env_measurements": work_env_count,
        "total_accidents": accident_count,
        "last_updated": datetime.now().isoformat()
    }