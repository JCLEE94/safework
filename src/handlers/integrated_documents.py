"""
통합문서관리 시스템 (Integrated Document Management System)
- 기존 PDF편집기 기능 통합
- PDF, Excel, Word 문서 편집 및 데이터 인서트
- 문서 템플릿 관리 및 양식 생성
"""

import io
import json
import logging
import os
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Union

# Office document libraries
import pandas as pd
from fastapi import (APIRouter, Depends, File, Form, HTTPException, Response,
                     UploadFile)
from fastapi.responses import FileResponse, StreamingResponse
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from pydantic import BaseModel, Field
from reportlab.lib.colors import black
import fitz  # PyMuPDF
from reportlab.lib.pagesizes import A4
from reportlab.pdfbase import pdfmetrics, pdfutils
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.pdfgen import canvas
from sqlalchemy.ext.asyncio import AsyncSession

from ..config.database import get_db
from ..config.settings import get_settings
from ..services.cache import CacheService, get_cache_service
from ..utils.auth_deps import get_current_user_id

# Initialize
router = APIRouter(prefix="/api/v1/integrated-documents", tags=["통합문서관리"])
logger = logging.getLogger(__name__)
settings = get_settings()

# Constants
DOCUMENT_BASE_DIR = Path("document")
UPLOAD_DIR = Path("uploads/documents")
TEMPLATE_DIR = Path("templates/documents")
SUPPORTED_FORMATS = {
    "pdf": [".pdf"],
    "excel": [".xlsx", ".xls"],
    "word": [".docx", ".doc"],
    "hwp": [".hwp"],
}

# Ensure directories exist
for directory in [UPLOAD_DIR, TEMPLATE_DIR]:
    directory.mkdir(parents=True, exist_ok=True)

# Document categories (from existing structure)
DOCUMENT_CATEGORIES = {
    "manual": {"name": "업무매뉴얼", "path": "01-업무매뉴얼"},
    "legal": {"name": "법정서식", "path": "02-법정서식"},
    "register": {"name": "관리대장", "path": "03-관리대장"},
    "checklist": {"name": "체크리스트", "path": "04-체크리스트"},
    "education": {"name": "교육자료", "path": "05-교육자료"},
    "msds": {"name": "MSDS관련", "path": "06-MSDS관련"},
    "special": {"name": "특별관리물질", "path": "07-특별관리물질"},
    "health": {"name": "건강관리", "path": "08-건강관리"},
    "reference": {"name": "참고자료", "path": "09-참고자료"},
    "latest": {"name": "최신자료", "path": "10-최신자료_2024-2025"},
}


# Pydantic models
class DocumentInfo(BaseModel):
    """문서 정보"""

    name: str
    path: str
    size: int
    category: str
    format: str
    created_at: datetime
    modified_at: datetime
    editable: bool = False


class DocumentEditRequest(BaseModel):
    """문서 편집 요청"""

    document_id: str
    edit_type: str = Field(..., description="edit_type: text, form, data")
    data: Dict[str, Any] = Field(default_factory=dict)
    position: Optional[Dict[str, float]] = None
    template_fields: Optional[Dict[str, str]] = None


class DocumentTemplate(BaseModel):
    """문서 템플릿"""

    name: str
    category: str
    format: str
    fields: List[Dict[str, Any]]
    layout: Dict[str, Any]


class ExcelEditRequest(BaseModel):
    """Excel 편집 요청"""

    sheet_name: Optional[str] = None
    cell_data: Dict[str, Any] = Field(default_factory=dict)
    range_data: Optional[Dict[str, List[List[Any]]]] = None
    style_options: Optional[Dict[str, Any]] = None


# Utility functions
def get_file_format(filename: str) -> str:
    """파일 확장자로 포맷 결정"""
    ext = Path(filename).suffix.lower()
    for format_name, extensions in SUPPORTED_FORMATS.items():
        if ext in extensions:
            return format_name
    return "unknown"


def is_editable_format(file_format: str) -> bool:
    """편집 가능한 포맷인지 확인"""
    return file_format in ["pdf", "excel", "word"]


async def scan_document_directory() -> List[DocumentInfo]:
    """문서 디렉토리 스캔"""
    documents = []

    for category_key, category_info in DOCUMENT_CATEGORIES.items():
        category_path = DOCUMENT_BASE_DIR / category_info["path"]
        if not category_path.exists():
            continue

        for file_path in category_path.rglob("*"):
            if file_path.is_file():
                stat = file_path.stat()
                file_format = get_file_format(file_path.name)

                if file_format != "unknown":
                    documents.append(
                        DocumentInfo(
                            name=file_path.name,
                            path=str(file_path.relative_to(DOCUMENT_BASE_DIR)),
                            size=stat.st_size,
                            category=category_key,
                            format=file_format,
                            created_at=datetime.fromtimestamp(stat.st_ctime),
                            modified_at=datetime.fromtimestamp(stat.st_mtime),
                            editable=is_editable_format(file_format),
                        )
                    )

    return documents


# PDF editing functions
def setup_korean_font():
    """한글 폰트 설정"""
    try:
        font_path = "/usr/share/fonts/truetype/nanum/NanumGothic.ttf"
        if os.path.exists(font_path):
            pdfmetrics.registerFont(TTFont("NanumGothic", font_path))
            return "NanumGothic"
        return "Helvetica"
    except Exception as e:
        logger.warning(f"한글 폰트 설정 실패: {e}")
        return "Helvetica"


def edit_pdf_with_data(pdf_path: Path, edit_data: Dict[str, Any]) -> bytes:
    """PDF에 데이터 삽입/편집"""
    try:
        # PDF 읽기
        pdf_doc = fitz.open(str(pdf_path))

        for page_num in range(len(pdf_doc)):
            page = pdf_doc[page_num]

            # 텍스트 추가
            if "text_inserts" in edit_data:
                for text_insert in edit_data["text_inserts"]:
                    rect = fitz.Rect(
                        text_insert["x"],
                        text_insert["y"],
                        text_insert["x"] + text_insert.get("width", 100),
                        text_insert["y"] + text_insert.get("height", 20),
                    )
                    page.insert_textbox(
                        rect,
                        text_insert["text"],
                        fontsize=text_insert.get("font_size", 12),
                        color=fitz.utils.getColor(text_insert.get("color", "black")),
                    )

            # 폼 필드 채우기
            if "form_fields" in edit_data:
                for field_name, field_value in edit_data["form_fields"].items():
                    widgets = page.widgets()
                    for widget in widgets:
                        if widget.field_name == field_name:
                            widget.field_value = str(field_value)
                            widget.update()

        # PDF 저장
        output_buffer = io.BytesIO()
        pdf_doc.save(output_buffer)
        pdf_doc.close()

        return output_buffer.getvalue()

    except Exception as e:
        logger.error(f"PDF 편집 오류: {e}")
        raise HTTPException(status_code=500, detail=f"PDF 편집 실패: {str(e)}")


# Excel editing functions
def edit_excel_with_data(excel_path: Path, edit_request: ExcelEditRequest) -> bytes:
    """Excel 파일 편집"""
    try:
        # Excel 파일 읽기
        if excel_path.suffix.lower() == ".xlsx":
            workbook = load_workbook(str(excel_path))
        else:
            # .xls 파일은 pandas로 읽고 openpyxl로 변환
            df = pd.read_excel(str(excel_path), sheet_name=None)
            workbook = Workbook()
            workbook.remove(workbook.active)

            for sheet_name, data in df.items():
                ws = workbook.create_sheet(sheet_name)
                for r_idx, row in enumerate(data.itertuples(index=False), 1):
                    for c_idx, value in enumerate(row, 1):
                        ws.cell(row=r_idx, column=c_idx, value=value)

        # 워크시트 선택
        if edit_request.sheet_name and edit_request.sheet_name in workbook.sheetnames:
            ws = workbook[edit_request.sheet_name]
        else:
            ws = workbook.active

        # 셀 데이터 편집
        if edit_request.cell_data:
            for cell_ref, value in edit_request.cell_data.items():
                ws[cell_ref] = value

        # 범위 데이터 편집
        if edit_request.range_data:
            for range_ref, data_matrix in edit_request.range_data.items():
                start_cell = ws[range_ref.split(":")[0]]
                start_row, start_col = start_cell.row, start_cell.column

                for row_idx, row_data in enumerate(data_matrix):
                    for col_idx, cell_value in enumerate(row_data):
                        ws.cell(
                            row=start_row + row_idx,
                            column=start_col + col_idx,
                            value=cell_value,
                        )

        # 스타일 적용
        if edit_request.style_options:
            style = edit_request.style_options
            if "font" in style:
                font = Font(**style["font"])
            if "fill" in style:
                fill = PatternFill(**style["fill"])
            if "alignment" in style:
                alignment = Alignment(**style["alignment"])

            # 스타일을 지정된 범위에 적용
            if "apply_range" in style:
                for row in ws[style["apply_range"]]:
                    for cell in row:
                        if "font" in style:
                            cell.font = font
                        if "fill" in style:
                            cell.fill = fill
                        if "alignment" in style:
                            cell.alignment = alignment

        # 파일 저장
        output_buffer = io.BytesIO()
        workbook.save(output_buffer)

        return output_buffer.getvalue()

    except Exception as e:
        logger.error(f"Excel 편집 오류: {e}")
        raise HTTPException(status_code=500, detail=f"Excel 편집 실패: {str(e)}")


# API endpoints
@router.get("/", response_model=List[DocumentInfo])
async def list_documents(
    category: Optional[str] = None,
    format: Optional[str] = None,
    editable_only: bool = False,
    cache: CacheService = Depends(get_cache_service),
):
    """전체 문서 목록 조회"""
    cache_key = f"documents:list:{category}:{format}:{editable_only}"
    cached_result = await cache.get(cache_key)

    if cached_result:
        return cached_result

    documents = await scan_document_directory()

    # 필터링
    if category:
        documents = [doc for doc in documents if doc.category == category]
    if format:
        documents = [doc for doc in documents if doc.format == format]
    if editable_only:
        documents = [doc for doc in documents if doc.editable]

    await cache.set(cache_key, documents, ttl=300)  # 5분 캐시
    return documents


@router.get("/categories")
async def get_categories():
    """문서 카테고리 목록"""
    return DOCUMENT_CATEGORIES


@router.get("/download/{category}/{path:path}")
async def download_document(category: str, path: str):
    """문서 다운로드"""
    if category not in DOCUMENT_CATEGORIES:
        raise HTTPException(status_code=404, detail="카테고리를 찾을 수 없습니다")

    file_path = DOCUMENT_BASE_DIR / DOCUMENT_CATEGORIES[category]["path"] / path

    if not file_path.exists():
        raise HTTPException(status_code=404, detail="파일을 찾을 수 없습니다")

    return FileResponse(
        path=str(file_path),
        filename=file_path.name,
        media_type="application/octet-stream",
    )


@router.post("/edit")
async def edit_document(
    edit_request: DocumentEditRequest, current_user_id=Depends(get_current_user_id)
):
    """문서 편집 (PDF/Excel)"""
    try:
        # 원본 파일 경로 구성
        document_path = None
        for category_key, category_info in DOCUMENT_CATEGORIES.items():
            potential_path = DOCUMENT_BASE_DIR / category_info["path"]
            for file_path in potential_path.rglob("*"):
                if file_path.name == edit_request.document_id or str(
                    file_path
                ).endswith(edit_request.document_id):
                    document_path = file_path
                    break
            if document_path:
                break

        if not document_path or not document_path.exists():
            raise HTTPException(status_code=404, detail="문서를 찾을 수 없습니다")

        file_format = get_file_format(document_path.name)

        # 편집 처리
        if file_format == "pdf":
            edited_content = edit_pdf_with_data(document_path, edit_request.data)
            media_type = "application/pdf"
            filename = f"edited_{document_path.stem}.pdf"

        elif file_format == "excel":
            excel_edit_req = ExcelEditRequest(**edit_request.data)
            edited_content = edit_excel_with_data(document_path, excel_edit_req)
            media_type = (
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            filename = f"edited_{document_path.stem}.xlsx"

        else:
            raise HTTPException(status_code=400, detail="지원하지 않는 파일 형식입니다")

        return StreamingResponse(
            io.BytesIO(edited_content),
            media_type=media_type,
            headers={"Content-Disposition": f"attachment; filename={filename}"},
        )

    except Exception as e:
        logger.error(f"문서 편집 오류: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/upload")
async def upload_document(
    file: UploadFile = File(...),
    category: str = Form(...),
    current_user_id=Depends(get_current_user_id),
):
    """새 문서 업로드"""
    if category not in DOCUMENT_CATEGORIES:
        raise HTTPException(status_code=400, detail="유효하지 않은 카테고리입니다")

    file_format = get_file_format(file.filename)
    if file_format == "unknown":
        raise HTTPException(status_code=400, detail="지원하지 않는 파일 형식입니다")

    # 업로드 디렉토리에 저장
    upload_path = UPLOAD_DIR / category
    upload_path.mkdir(parents=True, exist_ok=True)

    file_path = upload_path / file.filename

    with open(file_path, "wb") as buffer:
        content = await file.read()
        buffer.write(content)

    return {
        "message": "파일이 성공적으로 업로드되었습니다",
        "filename": file.filename,
        "category": category,
        "path": str(file_path.relative_to(UPLOAD_DIR)),
        "size": len(content),
    }


@router.delete("/{category}/{path:path}")
async def delete_document(
    category: str, path: str, current_user_id=Depends(get_current_user_id)
):
    """문서 삭제 (업로드된 파일만)"""
    if category not in DOCUMENT_CATEGORIES:
        raise HTTPException(status_code=404, detail="카테고리를 찾을 수 없습니다")

    # 업로드 디렉토리에서만 삭제 허용
    file_path = UPLOAD_DIR / category / path

    if not file_path.exists():
        raise HTTPException(status_code=404, detail="파일을 찾을 수 없습니다")

    file_path.unlink()

    return {"message": "파일이 삭제되었습니다"}


@router.get("/templates")
async def get_document_templates():
    """문서 템플릿 목록"""
    templates = []

    if TEMPLATE_DIR.exists():
        for template_file in TEMPLATE_DIR.rglob("*.json"):
            try:
                with open(template_file, "r", encoding="utf-8") as f:
                    template_data = json.load(f)
                    templates.append(DocumentTemplate(**template_data))
            except Exception as e:
                logger.warning(f"템플릿 로드 실패: {template_file} - {e}")

    return templates


@router.post("/templates")
async def create_document_template(
    template: DocumentTemplate, current_user_id=Depends(get_current_user_id)
):
    """새 문서 템플릿 생성"""
    template_file = TEMPLATE_DIR / f"{template.name}.json"

    with open(template_file, "w", encoding="utf-8") as f:
        json.dump(template.dict(), f, ensure_ascii=False, indent=2)

    return {"message": "템플릿이 생성되었습니다", "template": template}


@router.get("/stats")
async def get_document_stats(cache: CacheService = Depends(get_cache_service)):
    """문서 통계"""
    cache_key = "documents:stats"
    cached_stats = await cache.get(cache_key)

    if cached_stats:
        return cached_stats

    documents = await scan_document_directory()

    stats = {
        "total_documents": len(documents),
        "by_category": {},
        "by_format": {},
        "editable_count": sum(1 for doc in documents if doc.editable),
        "total_size": sum(doc.size for doc in documents),
    }

    for doc in documents:
        # 카테고리별 통계
        if doc.category not in stats["by_category"]:
            stats["by_category"][doc.category] = 0
        stats["by_category"][doc.category] += 1

        # 포맷별 통계
        if doc.format not in stats["by_format"]:
            stats["by_format"][doc.format] = 0
        stats["by_format"][doc.format] += 1

    await cache.set(cache_key, stats, ttl=600)  # 10분 캐시
    return stats
